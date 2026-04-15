"""
app/rag/chunker.py — Document Text Extraction & Chunking

Pipeline:
  Raw File → Text Extraction → Recursive Text Chunking → List[Chunk]

Supported formats:
- PDF (via pdfplumber for better accuracy)
- DOCX (via python-docx)
- TXT (plain text)
- XLSX/CSV (tabular data flattened to text)

Chunking strategy:
- Uses LangChain's RecursiveCharacterTextSplitter
- chunk_size=512 tokens (configurable)
- chunk_overlap=64 tokens (for context continuity)
- Splits on paragraphs → sentences → words (hierarchy)
"""

import os
import csv
from dataclasses import dataclass, field
from typing import List, Optional, Tuple
from langchain_text_splitters import RecursiveCharacterTextSplitter

from app.core.config import settings
from app.core.logging import get_logger

logger = get_logger(__name__)


@dataclass
class DocumentChunk:
    """Represents a single chunk of text from a document."""
    chunk_id: str        # Unique ID: f"{document_id}_chunk_{index}"
    document_id: int
    text: str
    chunk_index: int
    page_number: Optional[int] = None
    char_count: int = 0
    metadata: dict = field(default_factory=dict)

    def __post_init__(self):
        self.char_count = len(self.text)


# ─── Text Extraction ──────────────────────────────────────────────────────────

def extract_text_from_pdf(file_path: str) -> List[Tuple[str, int]]:
    """
    Extract text from PDF with page numbers.

    Returns:
        List of (text, page_number) tuples
    """
    try:
        import pdfplumber
        pages = []
        with pdfplumber.open(file_path) as pdf:
            for i, page in enumerate(pdf.pages, start=1):
                text = page.extract_text()
                if text and text.strip():
                    pages.append((text.strip(), i))
        return pages
    except Exception as e:
        logger.error(f"PDF extraction error for {file_path}: {e}")
        return []


def extract_text_from_docx(file_path: str) -> List[Tuple[str, int]]:
    """
    Extract text from DOCX.
    Returns (full_text, 0) since DOCX doesn't have page numbers.
    """
    try:
        from docx import Document
        doc = Document(file_path)
        paragraphs = [p.text for p in doc.paragraphs if p.text.strip()]
        full_text = "\n\n".join(paragraphs)
        return [(full_text, 0)]
    except Exception as e:
        logger.error(f"DOCX extraction error for {file_path}: {e}")
        return []


def extract_text_from_txt(file_path: str) -> List[Tuple[str, int]]:
    """Extract text from plain text files."""
    try:
        with open(file_path, "r", encoding="utf-8", errors="replace") as f:
            content = f.read()
        return [(content, 0)]
    except Exception as e:
        logger.error(f"TXT extraction error for {file_path}: {e}")
        return []


def extract_text_from_xlsx(file_path: str) -> List[Tuple[str, int]]:
    """
    Extract text from Excel files.
    Converts each sheet to a text representation.
    """
    try:
        import openpyxl
        wb = openpyxl.load_workbook(file_path, data_only=True)
        sheets_text = []
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            rows = []
            for row in ws.iter_rows(values_only=True):
                row_text = " | ".join(str(cell) for cell in row if cell is not None)
                if row_text.strip():
                    rows.append(row_text)
            if rows:
                sheet_text = f"Sheet: {sheet_name}\n" + "\n".join(rows)
                sheets_text.append(sheet_text)
        full_text = "\n\n".join(sheets_text)
        return [(full_text, 0)]
    except Exception as e:
        logger.error(f"XLSX extraction error for {file_path}: {e}")
        return []


def extract_text_from_csv(file_path: str) -> List[Tuple[str, int]]:
    """Extract text from CSV files."""
    try:
        rows = []
        with open(file_path, "r", encoding="utf-8", errors="replace") as f:
            reader = csv.reader(f)
            for row in reader:
                row_text = " | ".join(str(cell) for cell in row if cell)
                if row_text.strip():
                    rows.append(row_text)
        return [("\n".join(rows), 0)]
    except Exception as e:
        logger.error(f"CSV extraction error for {file_path}: {e}")
        return []


def extract_text(file_path: str) -> List[Tuple[str, int]]:
    """
    Route text extraction based on file extension.

    Args:
        file_path: Full path to the file on disk

    Returns:
        List of (text_content, page_number) tuples

    Raises:
        ValueError: If file extension is not supported
    """
    ext = os.path.splitext(file_path)[1].lower()
    extractors = {
        ".pdf": extract_text_from_pdf,
        ".docx": extract_text_from_docx,
        ".txt": extract_text_from_txt,
        ".xlsx": extract_text_from_xlsx,
        ".csv": extract_text_from_csv,
    }

    extractor = extractors.get(ext)
    if not extractor:
        raise ValueError(f"Unsupported file extension: {ext}")

    return extractor(file_path)


# ─── Text Chunking ────────────────────────────────────────────────────────────

def create_chunks(
    document_id: int,
    file_path: str,
    metadata: dict = None,
    chunk_size: int = None,
    chunk_overlap: int = None,
) -> List[DocumentChunk]:
    """
    Full pipeline: Extract text from file → split into chunks.

    LangChain's RecursiveCharacterTextSplitter splits on:
    ["\n\n", "\n", ". ", " ", ""] — in order of preference
    This preserves paragraph and sentence boundaries.

    Args:
        document_id: Database ID of the document
        file_path: Path to the file on disk
        metadata: Additional metadata to attach to each chunk
        chunk_size: Override default chunk size from settings
        chunk_overlap: Override default overlap from settings

    Returns:
        List of DocumentChunk objects ready for embedding
    """
    chunk_size = chunk_size or settings.CHUNK_SIZE
    chunk_overlap = chunk_overlap or settings.CHUNK_OVERLAP
    metadata = metadata or {}

    # Step 1: Extract raw text with page numbers
    pages = extract_text(file_path)
    if not pages:
        logger.warning(f"No text extracted from document {document_id}: {file_path}")
        return []

    logger.info(f"Extracted {len(pages)} pages from document {document_id}")

    # Step 2: Initialize LangChain text splitter
    splitter = RecursiveCharacterTextSplitter(
        chunk_size=chunk_size,
        chunk_overlap=chunk_overlap,
        length_function=len,
        separators=["\n\n", "\n", ". ", " ", ""],
        is_separator_regex=False,
    )

    # Step 3: Chunk each page/section separately
    chunks = []
    chunk_index = 0

    for page_text, page_number in pages:
        # Split this page's text into sub-chunks
        page_chunks = splitter.split_text(page_text)

        for text in page_chunks:
            text = text.strip()
            if not text:
                continue

            chunk = DocumentChunk(
                chunk_id=f"doc_{document_id}_chunk_{chunk_index}",
                document_id=document_id,
                text=text,
                chunk_index=chunk_index,
                page_number=page_number if page_number > 0 else None,
                metadata={
                    **metadata,
                    "page_number": page_number,
                    "document_id": str(document_id),
                },
            )
            chunks.append(chunk)
            chunk_index += 1

    logger.info(
        f"Document {document_id} split into {len(chunks)} chunks "
        f"(chunk_size={chunk_size}, overlap={chunk_overlap})"
    )
    return chunks
